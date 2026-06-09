import argparse
import os
import re
import psycopg2
import sys
import unicodedata
from datetime import datetime

from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from database import get_db_connection

#DEFAULT_EXCEL_PATH = r"C:\Users\Carlo\Desktop\Base de Prueba.xlsx"
DEFAULT_EXCEL_PATH = r"C:\Users\ipsd4\Desktop\Base de Prueba.xlsx"

def normalizar_texto(valor):
    return str(valor or '').strip()

def normalizar_correo(correo):
    return normalizar_texto(correo).lower()

def normalizar_cabecera(cabecera):
    texto = normalizar_texto(cabecera)
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    texto = re.sub(r'\s+', ' ', texto).strip().lower()
    return texto

def validar_numero_empleado(numero):
    return bool(re.match(r'^\d{3,12}$', numero))

def validar_correo(correo):
    return bool(re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', correo))

def asegurar_tabla_docentes(conn):
    with conn.cursor() as cursor:
        cursor.execute(
            '''
            CREATE TABLE IF NOT EXISTS docentes (
                id SERIAL PRIMARY KEY,
                numero_empleado TEXT UNIQUE NOT NULL,
                nombre_completo TEXT NOT NULL,
                correo_institucional TEXT UNIQUE NOT NULL,
                centro_universitario_regional TEXT NOT NULL DEFAULT '',
                activo INTEGER NOT NULL DEFAULT 1,
                fecha_sincronizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            '''
        )
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_docentes_numero ON docentes (numero_empleado)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_docentes_correo ON docentes (correo_institucional)')
    conn.commit()

def resolver_columnas(headers):
    columnas_normalizadas = [normalizar_cabecera(c) for c in headers]

    alias_map = {
        'numero_empleado': {
            'numero de empleado',
            'numero empleado',
            'n de empleado',
            'no empleado',
            'num empleado',
        },
        'nombre_completo': {
            'nombre completo',
            'nombre',
        },
        'correo_institucional': {
            'correo institucional',
            'correo',
            'email institucional',
            'correo electronico',
            'correo electronico institucional',
        },
        'centro_universitario_regional': {
            'centro universitario regional',
            'centro universitario',
            'centro regional',
            'cur',
        },
    }

    indices = {}
    for campo, alias in alias_map.items():
        for idx, nombre_col in enumerate(columnas_normalizadas):
            if nombre_col in alias:
                indices[campo] = idx
                break

    obligatorias = {'numero_empleado', 'nombre_completo', 'correo_institucional'}
    faltantes = [k for k in obligatorias if k not in indices]
    if faltantes:
        raise ValueError(
            'No se encontraron columnas obligatorias en el Excel: ' + ', '.join(faltantes)
        )

    return indices

def sincronizar_docentes(excel_path, sheet_name=None, desactivar_ausentes=True):
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f'La hoja "{sheet_name}" no existe en el archivo Excel.')
        hoja = wb[sheet_name]
    else:
        hoja = wb.active

    if hoja is None:
        wb.close()
        raise ValueError('No se pudo determinar la hoja a procesar en el archivo Excel.')

    rows_iter = hoja.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration as exc:
        wb.close()
        raise ValueError('El archivo Excel no contiene filas para procesar.') from exc

    indices = resolver_columnas(headers)

    total_filas = 0
    registros_validos = []
    errores_validacion = 0

    for row in rows_iter:
        total_filas += 1
        numero_empleado = normalizar_texto(row[indices['numero_empleado']] if indices['numero_empleado'] < len(row) else '')
        nombre_completo = normalizar_texto(row[indices['nombre_completo']] if indices['nombre_completo'] < len(row) else '')
        correo_institucional = normalizar_correo(
            row[indices['correo_institucional']] if indices['correo_institucional'] < len(row) else ''
        )
        centro_universitario_regional = ''
        if 'centro_universitario_regional' in indices:
            centro_universitario_regional = normalizar_texto(
                row[indices['centro_universitario_regional']] if indices['centro_universitario_regional'] < len(row) else ''
            )

        if not numero_empleado or not nombre_completo or not correo_institucional:
            print(f"[ERROR VALIDACIÓN] Fila {total_filas + 1}: Faltan campos obligatorios. Empleado: '{numero_empleado}', Nombre: '{nombre_completo}', Correo: '{correo_institucional}'")
            errores_validacion += 1
            continue

        if not validar_numero_empleado(numero_empleado):
            print(f"[ERROR VALIDACIÓN] Fila {total_filas + 1}: Número de empleado inválido '{numero_empleado}' (Debe ser numérico de 3 a 12 dígitos)")
            errores_validacion += 1
            continue

        if not validar_correo(correo_institucional):
            print(f"[ERROR VALIDACIÓN] Fila {total_filas + 1}: Formato de correo institucional inválido '{correo_institucional}'")
            errores_validacion += 1
            continue

        registros_validos.append(
            (numero_empleado, nombre_completo, correo_institucional, centro_universitario_regional)
        )

    wb.close()

    conn = get_db_connection()
    try:
        asegurar_tabla_docentes(conn)
        
        with conn.cursor() as cursor:
            if desactivar_ausentes:
                cursor.execute('UPDATE docentes SET activo = 0')

        upserts_ok = 0
        conflictos = 0

        for numero_empleado, nombre_completo, correo_institucional, centro_universitario_regional in registros_validos:
            try:
                with conn.cursor() as cursor:
                    cursor.execute(
                        '''
                        INSERT INTO docentes (
                            numero_empleado, nombre_completo, correo_institucional, centro_universitario_regional, activo, fecha_sincronizacion
                        )
                        VALUES (%s, %s, %s, %s, 1, CURRENT_TIMESTAMP)
                        ON CONFLICT(numero_empleado) DO UPDATE SET
                            nombre_completo = EXCLUDED.nombre_completo,
                            correo_institucional = EXCLUDED.correo_institucional,
                            centro_universitario_regional = EXCLUDED.centro_universitario_regional,
                            activo = 1,
                            fecha_sincronizacion = CURRENT_TIMESTAMP
                        ''',
                        (numero_empleado, nombre_completo, correo_institucional, centro_universitario_regional),
                    )
                upserts_ok += 1
                # Hacemos commit por cada registro exitoso para no perder el progreso si uno falla
                conn.commit()
            except psycopg2.IntegrityError:
                # Si falla (ej. correo duplicado), Postgres bloquea la transacción. Esto la resetea.
                conn.rollback()
                conflictos += 1

    finally:
        conn.close()

    # Usamos os.environ para que no dé error si quieres imprimir el nombre de la DB
    db_nombre = os.environ.get('DATABASE_URL', 'PostgreSQL DB')

    return {
        'archivo': excel_path,
        'hoja': hoja.title,
        'db_path': db_nombre,
        'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_filas': total_filas,
        'validos': len(registros_validos),
        'errores_validacion': errores_validacion,
        'upserts_ok': upserts_ok,
        'conflictos': conflictos,
        'desactivar_ausentes': desactivar_ausentes,
    }

def parse_args():
    parser = argparse.ArgumentParser(
        description='Sincroniza manualmente docentes desde un archivo Excel hacia PostgreSQL.'
    )
    parser.add_argument(
        '--excel',
        default=DEFAULT_EXCEL_PATH,
        help='Ruta del archivo Excel con datos de docentes.',
    )
    parser.add_argument(
        '--sheet',
        default=None,
        help='Nombre de la hoja a procesar. Si se omite, usa la hoja activa.',
    )
    parser.add_argument(
        '--no-desactivar-ausentes',
        action='store_true',
        help='Si se usa, no desactiva docentes ausentes en el Excel.',
    )
    return parser.parse_args()

def main():
    args = parse_args()
    resultado = sincronizar_docentes(
        excel_path=args.excel,
        sheet_name=args.sheet,
        desactivar_ausentes=not args.no_desactivar_ausentes,
    )

    print('Sincronizacion completada')
    print(f"Fecha: {resultado['fecha']}")
    print(f"Archivo: {resultado['archivo']}")
    print(f"Hoja: {resultado['hoja']}")
    print(f"Conexion: {resultado['db_path']}")
    print(f"Filas leidas: {resultado['total_filas']}")
    print(f"Registros validos: {resultado['validos']}")
    print(f"Registros sincronizados (insert/update): {resultado['upserts_ok']}")
    print(f"Errores de validacion: {resultado['errores_validacion']}")
    print(f"Conflictos por restricciones unicas: {resultado['conflictos']}")
    print(f"Docentes ausentes desactivados: {'si' if resultado['desactivar_ausentes'] else 'no'}")

if __name__ == '__main__':
    main()